#!/usr/bin/env python3


# This script is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This script is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Lesser General Public License for more details.

# You should have received a copy of the GNU Lesser General Public License
# along with this script. If not, see <https://www.gnu.org/licenses/>.

# Author: Zophonías Oddur Jónsson (with assistant from Copilot)


"""
Pivot exam grades, optionally merge with student info, and format Excel output.

Usage:
    ./pivot_grades.py grades.csv [-s STUDENTS.xlsx] [-o OUTPUT.xlsx] [-c ORDERFILE.txt] [-r REGEX] [--dry-run]

Options:
    grades.csv            Path to grades CSV file
    -s, --students        Optional student info Excel file
    -o, --out             Optional output Excel filename
    -c, --columnorder     Optional column order file (one QuestionTitle per line)
    -r, --regex           Optional regex for sorting question columns (must have one capturing group)
    --help-regex          Show detailed help on regex usage and exit
    --dry-run             Show planned column order and exit without writing Excel file

Regex Notes:
    - The regex must contain ONE capturing group.
    - The captured value will be used as the sort key.
    - If the captured value is numeric, sorting will be numeric; otherwise, lexicographic.

Example:
    ./pivot_grades.py grades.csv -r "Okt24-(\\d+)"
    This extracts the number after 'Okt24-' and sorts questions numerically.
"""

import pandas as pd
import os
import argparse
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def print_regex_help():
    print("""
Regex Sorting Help:
--------------------
Use -r or --regex to define how question columns should be sorted.
The regex must contain ONE capturing group. The captured value is used as the sort key.
If numeric, sorting is numeric; otherwise, lexicographic.

Examples:
  -r "Okt24-(\\d+)"   # Sorts by the number after 'Okt24-'
  -r "Q(\\d+)"        # Sorts Q1, Q2, Q10 numerically
  -r "([A-Za-z]+)"      # Sorts by alphabetic prefix

Tips:
- Always escape backslashes in the shell: use \\\\d+ instead of \\d+.
- If regex doesn't match a column, that column falls back to its original name.
""")
    exit(0)

def load_grades(grades_file):
    df = pd.read_csv(grades_file, sep=None, engine='python')
    df['FinalScore'] = df['ManuallyGradedScore'].fillna(df['AutoGradedScore'])
    return df

def create_pivot(df, order_file=None, regex=None):
    pivot_df = df.pivot_table(
        index=['CandidateExternalId', 'UserId'],
        columns='QuestionTitle',
        values='FinalScore',
        aggfunc='first'
    )

    if order_file and regex:
        raise ValueError("Cannot use both column order file (-c) and regex (-r). Choose one.")

    if order_file:
        with open(order_file, 'r', encoding='utf-8') as f:
            desired_order = [line.strip() for line in f if line.strip()]
        ordered_cols = [col for col in desired_order if col in pivot_df.columns]
        remaining_cols = [col for col in pivot_df.columns if col not in ordered_cols]
        final_cols = ordered_cols + remaining_cols
    elif regex:
        pattern = re.compile(regex)
 #       def sort_key(col):
 #           match = pattern.search(col)
 #           if match:
 #               val = match.group(1)
 #               return int(val) if val.isdigit() else val
 #           return col
 
        def sort_key(col):
            match = re.match(regex, str(col))
            if match:
                return int(match.group(1))  # Skilar heiltölu
            return float('inf')  # Setur óraðanlegar strengi aftast
        final_cols = sorted(pivot_df.columns, key=sort_key)
    else:
        final_cols = sorted(pivot_df.columns)

    pivot_df = pivot_df[final_cols]
    pivot_df.reset_index(inplace=True)
    return pivot_df, final_cols

def load_students(students_file):
    return pd.read_excel(students_file)

def merge_data(students_df, pivot_df):
    return pd.merge(
        students_df,
        pivot_df,
        how='outer',
        left_on='Nafn í prófi',
        right_on='CandidateExternalId'
    )

def write_excel(df, output_file):
    df.to_excel(output_file, index=False)

def apply_excel_formatting(output_file, question_cols):
    wb = load_workbook(output_file)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    header_row = [cell.value for cell in ws[1]]
    question_indexes = [header_row.index(q) + 1 for q in question_cols if q in header_row]

    if question_indexes:
        first_q_col = question_indexes[0]
        last_q_col = question_indexes[-1]
        total_col = ws.max_column + 1
        ws.cell(row=1, column=total_col, value="Total Score")

        for row in range(2, ws.max_row + 1):
            first_letter = get_column_letter(first_q_col)
            last_letter = get_column_letter(last_q_col)
            ws.cell(row=row, column=total_col, value=f"=SUM({first_letter}{row}:{last_letter}{row})")

    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(
        description="Pivot grades and optionally merge with student info.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("grades", nargs="?", help="Path to grades CSV file")
    parser.add_argument("-s", "--students", help="Optional student info Excel file")
    parser.add_argument("-o", "--out", help="Output Excel filename")
    parser.add_argument("-c", "--columnorder", help="Optional column order file")
    parser.add_argument("-r", "--regex", help="Optional regex for sorting question columns (must have one capturing group)")
    parser.add_argument("--help-regex", action="store_true", help="Show detailed help on regex usage and exit")
    parser.add_argument("--dry-run", action="store_true", help="Show planned column order and exit without writing Excel file")

    args = parser.parse_args()

    if args.help_regex:
        print_regex_help()

    if not args.grades:
        parser.print_help()
        exit(1)

    if args.columnorder and args.regex:
        print("Error: Cannot use both -c (column order file) and -r (regex).")
        exit(1)

    grades_df = load_grades(args.grades)
    pivot_df, question_cols = create_pivot(grades_df, args.columnorder, args.regex)

    if args.dry_run:
        print("\nPlanned column order:")
        for col in question_cols:
            print(f"  {col}")
        print("\n(Dry run: no file written.)")
        exit(0)

    if args.out:
        output_file = args.out
    else:
        dirname, basename = os.path.split(args.grades)
        name, _ = os.path.splitext(basename)
        output_file = os.path.join(dirname, f"pivoted-{name}.xlsx")

    if args.students:
        students_df = load_students(args.students)
        final_df = merge_data(students_df, pivot_df)
    else:
        final_df = pivot_df

    write_excel(final_df, output_file)
    apply_excel_formatting(output_file, question_cols)

    print(f"✅ File saved to {output_file}")

if __name__ == "__main__":
    main()
